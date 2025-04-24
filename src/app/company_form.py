import streamlit as st
import pdfplumber
from docx import Document
from openpyxl import load_workbook
import pandas as pd
import json
import tempfile
import os
from typing import Dict, List, Tuple
from collections import defaultdict

# Set page configuration (MUST be first Streamlit command)
st.set_page_config(page_title="GHG Emissions Guidance Form", layout="wide")


def display_company_form():
    # Define default values
    defaults = {
        "company_name": "",
        "industry": "",
        "revenue_range": "",
        "company_size": "",
        "reporting_status": "",
        "emission_scopes": [],
        "emission_sources": [],
        "objective": "",
        "output_format": "",
        "regulatory_basis": "",
        "reference_support": False,
        "challenges": "",
        "submitted": False,
        "fleet_size": None,
        "electricity": None
    }

    # Initialize session state
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val

    # App Header
    st.header("GHG Emissions Guidance Form")
    st.text(
        "Fill out this form to receive tailored guidance based on Australia's 2025 regulations or international frameworks."
    )

    # Section 1: Company Info
    st.subheader("1. Company Information")
    col1, col2 = st.columns(2)
    with col1:
        company_name = st.text_input(
            "Company Name (Optional)", value=st.session_state.company_name
        )
        industry = st.selectbox(
            "Industry/Sector*",
            [
                "",
                "Mining",
                "Manufacturing",
                "Agriculture",
                "Energy",
                "Transport",
                "Construction",
                "Retail",
                "Other",
            ],
            index=(
                0
                if st.session_state.industry == ""
                else [
                    "",
                    "Mining",
                    "Manufacturing",
                    "Agriculture",
                    "Energy",
                    "Transport",
                    "Construction",
                    "Retail",
                    "Other",
                ].index(st.session_state.industry)
            ),
        )
        revenue_range = st.selectbox(
            "Annual Revenue (AUD)",
            ["", "<50M", "50M–200M", "200M–500M", ">500M", "Prefer not to say"],
            index=(
                0
                if st.session_state.revenue_range == ""
                else [
                    "",
                    "<50M",
                    "50M–200M",
                    "200M–500M",
                    ">500M",
                    "Prefer not to say",
                ].index(st.session_state.revenue_range)
            ),
        )
    with col2:
        company_size = st.selectbox(
            "Company Size*",
            ["", "Small (<20 employees)", "Medium (20–199)", "Large (200+)"],
            index=(
                0
                if st.session_state.company_size == ""
                else [
                    "",
                    "Small (<20 employees)",
                    "Medium (20–199)",
                    "Large (200+)",
                ].index(st.session_state.company_size)
            ),
        )
        location = st.selectbox(
              "Location (State/Territory)",
             ["", "NSW", "VIC", "QLD", "WA", "SA", "TAS", "ACT", "NT"]
        )

    # Section 2: Emissions Profile
    st.subheader("2. Emissions Profile")
    reporting_status = st.radio(
        "Current GHG Reporting Status*",
        ["Already reporting", "New to reporting", "Unsure"],
        index=(
            ["Already reporting", "New to reporting", "Unsure"].index(
                st.session_state.reporting_status
            )
            if st.session_state.reporting_status
            else 0
        ),
        horizontal=True,
    )

    emission_scopes = st.multiselect(
        "Which emission scopes are you assessing?*",
        [
            "Scope 1 (Direct)",
            "Scope 2 (Indirect – Energy)",
            "Scope 3 (Value chain)",
            "Unsure – need guidance",
        ],
        default=st.session_state.emission_scopes,
    )

    current_sources = st.multiselect(
        "Key Emission Sources",
        [
            "Stationary combustion (boilers, generators)",
            "Mobile combustion (fleets, transport)",
            "Electricity use",
            "Fugitive emissions (e.g. refrigerants, methane)",
            "Waste and wastewater",
            "Purchased goods and services",
            "Employee commuting / travel",
            "Use of sold products",
            "Other Scope 3 sources",
        ],
        default=st.session_state.emission_sources,
    )

    # Section 3: Operational Details
    st.subheader("3. Operational Details")
    electricity = None
    renewables = None
    fleet_size = None
    fuel_type = None

    if "Electricity use" in current_sources:
        st.markdown("**Energy Usage**")
        electricity = st.text_input(
            "Annual Electricity Consumption (kWh)", placeholder="Please insert the Annual Electricity Consumption in kWh", 
            value=st.session_state.electricity
        )
        renewables = st.slider("Renewable Energy Usage (%)", 0, 100, 0)

    if "Mobile combustion (fleets, transport)" in current_sources:
        st.markdown("**Fleet / Transport**")
        fleet_size = st.text_input("Number of Vehicles", placeholder="Please insert the number of vehicles",
                                   value= st.session_state.fleet_size)
        fuel_type = st.selectbox(
            "Primary Fuel Type", ["Petrol", "Diesel", "Electric", "Hybrid"]
        )

    # Section 4: Goals
    st.subheader("4. Goals & Preferences")
    objective = st.selectbox(
        "Primary Objective",
        ["Compliance only", "Reduction strategy", "Net-zero planning", "Educational"],
        index=(
            0
            if st.session_state.objective == ""
            else [
                "Compliance only",
                "Reduction strategy",
                "Net-zero planning",
                "Educational",
            ].index(st.session_state.objective)
        ),
    )

    output_format = st.radio(
        "Preferred Output Format",
        ["Step-by-step guide", "Summary report", "Regulatory citations"],
        index=(
            ["Step-by-step guide", "Summary report", "Regulatory citations"].index(
                st.session_state.output_format
            )
            if st.session_state.output_format
            else 0
        ),
        horizontal=True,
    )

    # Section 5: Framework Reference
    st.subheader("5. Standards & Framework Preferences")
    regulatory_basis = st.selectbox(
        "Which regulatory framework should the guidance follow?",
        [
            "Australia 2025 Climate Reporting (Default)",
            "United States – EPA Framework",
            "European Union – ESRS / CSRD",
            "General / International Guidance (GHG Protocol, ISO, API)",
        ],
        index=(
            0
            if st.session_state.regulatory_basis == ""
            else [
                "Australia 2025 Climate Reporting (Default)",
                "United States – EPA Framework",
                "European Union – ESRS / CSRD",
                "General / International Guidance (GHG Protocol, ISO, API)",
            ].index(st.session_state.regulatory_basis)
        ),
    )

    reference_support = st.checkbox(
        "Include references to international best practices (GHG Protocol, ISO 14064/67, API Compendium, etc.)",
        value=st.session_state.reference_support,
    )

    st.subheader("(Optional) Additional Info")
    challenges = st.text_area(
        "Any specific challenges you’re facing?", value=st.session_state.challenges
    )

    uploaded_files = st.file_uploader(
        "Upload Additional documents (PDF, DOCX, XLSX, CSV, TXT, or JSON)",
        type=["pdf", "docx", "xlsx", "csv", "txt", "json"],
        accept_multiple_files=True,
    )

    if uploaded_files:
        if "processed_files" not in st.session_state:
            st.session_state.processed_files = defaultdict(dict)

        with st.spinner(f"Processing {len(uploaded_files)} files..."):
            for file in uploaded_files:
                if file.name in st.session_state.processed_files:
                    continue

                filename, text = process_single_file(file)
                if not text.strip():
                    st.session_state.processed_files[filename] = {
                        "status": "failed",
                        "message": "No text content found",
                    }
                    continue

                embeddings = (
                    st.session_state.rag_class.embedding_class.custom_embeddings([text])
                )
                st.session_state.processed_files[filename] = {"embeddings": embeddings}

    # Submit Button
    if st.button("Submit Form"):
        valid_form = True
        if not industry or not company_size or not emission_scopes:
            valid_form = False
            st.error("Please fill required fields (*)")

        if "Electricity use" in current_sources and electricity is not None:
            try:
                electricity = float(electricity)
            except:
                valid_form = False
                st.error("Annual Electricity Consumption needs to be a number.")
        
        if "Mobile combustion (fleets, transport)" in current_sources and fleet_size is not None:
            try:
                fleet_size = int(fleet_size)
            except:
                valid_form = False
                st.error("Fleet size needs to be a number.")

        if valid_form:
            # Save to session state
            st.session_state.submitted = True
            st.session_state.company_name = company_name
            st.session_state.industry = industry
            st.session_state.revenue_range = revenue_range
            st.session_state.company_size = company_size
            st.session_state.reporting_status = reporting_status
            st.session_state.emission_scopes = emission_scopes
            st.session_state.emission_sources = current_sources
            st.session_state.objective = objective
            st.session_state.output_format = output_format
            st.session_state.regulatory_basis = regulatory_basis
            st.session_state.reference_support = reference_support
            st.session_state.challenges = challenges
            st.session_state.fleet_size = fleet_size
            st.session_state.electricity = electricity

            # Prepare operational data for LLM context
            operational_details = ""
            if "Electricity use" in current_sources and electricity is not None:
                operational_details += (
                    f"- Annual Electricity Consumption: {electricity} kWh\n"
                )
                operational_details += f"- Renewable Energy Usage: {renewables}%\n"
            if (
                "Mobile combustion (fleets, transport)" in current_sources
                and fleet_size is not None
            ):
                operational_details += f"- Fleet Size: {fleet_size} vehicles\n"
                operational_details += f"- Fuel Type: {fuel_type}\n"

            # LLM Context
            context = f"""
The company operates in Australia** and is subject to Australia's evolving climate reporting regulations, effective 2025.

The user has selected **{regulatory_basis}** as the preferred framework for guidance.
{"Please include reference section at the end with the specific document and framework use URL links for documents." if reference_support else "Do not include reference section at the end.."}

Company Info:
- Company Name: {company_name}
- Industry: {industry}
- Company Size: {company_size}
- Revenue Range: {revenue_range}
- Reporting Status: {reporting_status}
- Scopes Being Assessed: {", ".join(emission_scopes)}
- Key Emission Sources: {", ".join(current_sources)}
- Objective: {objective}
- Preferred Output Format: {output_format}
- Location: {location}

Challenges: {challenges if challenges else "None provided"}

Operational Details:
{operational_details if operational_details else "Not provided"}
"""
            if uploaded_files:
                st.session_state.ghg_assistant.set_context_form(context, st.session_state.processed_files)
            else:
                 st.session_state.ghg_assistant.set_context_form(context)
            st.success("Form submitted! Guidance is being prepared.")
            st.balloons()


def extract_text_from_pdf(file):
    text = ""
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def extract_text_from_docx(file):
    doc = Document(file)
    return "\n".join([paragraph.text for paragraph in doc.paragraphs])


def extract_text_from_xlsx(file):
    text = ""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(file.read())
        tmp_path = tmp.name

    try:
        wb = load_workbook(tmp_path)
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                if row_text:
                    text += row_text + "\n"
    finally:
        os.unlink(tmp_path)
    return text


def extract_text_from_csv(file):
    df = pd.read_csv(file)
    return "\n".join(df.astype(str).apply(lambda row: " ".join(row), axis=1))


def extract_text_from_txt(file):
    return file.read().decode("utf-8")


def extract_text_from_json(file):
    data = json.load(file)
    if isinstance(data, list):
        return " ".join(str(item) for item in data)
    elif isinstance(data, dict):
        return " ".join(f"{key}: {value}" for key, value in data.items())
    else:
        return str(data)


def process_single_file(file) -> Tuple[str, str]:
    file_type = file.name.split(".")[-1].lower()

    try:
        if file_type == "pdf":
            return file.name, extract_text_from_pdf(file)
        elif file_type == "docx":
            return file.name, extract_text_from_docx(file)
        elif file_type == "xlsx":
            return file.name, extract_text_from_xlsx(file)
        elif file_type == "csv":
            return file.name, extract_text_from_csv(file)
        elif file_type == "txt":
            return file.name, extract_text_from_txt(file)
        elif file_type == "json":
            return file.name, extract_text_from_json(file)
        else:
            return file.name, ""
    except Exception as e:
        st.error(f"Error processing {file.name}: {str(e)}")
        return file.name, ""
